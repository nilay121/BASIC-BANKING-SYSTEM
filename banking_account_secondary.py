###################BASIC BANKING SYSTEM
###################PRIMARY ACCOUNT
##################MAIN ACCOUNT WE NEED TO RUN THIS IF WE WANT TO SEND    MONEY
##################DESIGNED BY NILAY KUSHAWAHA

from banking_account_primary import account_details

def sent_money_primary():
    import openpyxl as xl
    wb = xl.load_workbook('account_details_updated.xlsx')
    sheet = wb['Sheet1']
    amount_received = account_details()
    received = sheet.cell(3, 5)
    received.value = amount_received
    money2 = sheet.cell(3, 4)
    amount = money2.value
    new_amount2 = amount + amount_received
    money2.value = new_amount2
    wb.save('account_details_updated.xlsx')
    print("##### Money sent and received #####")

sent_money_primary()