###################BASIC BANKING SYSTEM
###################SECONDARY ACCOUNT
##################DESIGNED BY NILAY KUSHAWAHA
def account_details():
    import openpyxl as xl
    wb = xl.load_workbook('account_details_updated.xlsx')
    sheet = wb['Sheet1']
    name = input("Enter your name as registered in bank :- ")
    while(True):
        print(f"Welcome {name.upper()}\nDo you want to continue with the account detils? \nType y or n for yes or no")
        received = input()
        if received.upper() == 'Y':
                stage2=int(input(" 1. NEW ACCOUNT \n 2.ACCOUNT DETAILS\n"))
                if stage2 == 1:
                    data = sheet.cell(3,1)
                    name_taken = input("Enter your Name :- ")
                    data.value = name_taken
                    break
                elif stage2 == 2:
                    option = int(input("\n1.Amount Available \n2.Deposit\n3.Withdraw\n4.Send\n"))
                    if option == 1:
                        money = sheet.cell(2,4)
                        amount = money.value
                        print(amount)
                        return 0
                        break
                    elif option == 2:
                        addition = int(input("Enter the amount you want to add :- "))
                        deposited = sheet.cell(2,2)
                        deposited.value = addition
                        money = sheet.cell(2, 4)
                        amount = money.value
                        print(f"Amount {addition} has been added to the account ")
                        new_amount = addition + amount
                        print(new_amount)
                        money.value = new_amount
                        return 0
                        break
                    elif option == 3:
                        subtracted = int(input("Enter the amount you want to add :- "))
                        withdrawl = sheet.cell(2,3)
                        withdrawl.value = subtracted
                        money = sheet.cell(2, 4)
                        amount = money.value
                        print(f"Amount {subtracted} has been deducted from your account ")
                        new_amount = amount - subtracted
                        print(new_amount)
                        money.value = new_amount
                        return 0
                        break
                    elif option == 4:
                        sent_user = int(input("Enter the amount to send :- "))
                        sent = sheet.cell(2, 5)
                        sent.value = sent_user
                        money = sheet.cell(2, 4)
                        amount = money.value
                        print(f"Amount {sent_user} has been deducted from your account ")
                        new_amount = amount - sent_user
                        print(new_amount)
                        money.value = new_amount
                        return sent_user
                        break
        elif received.upper() == 'N':
            print(f"Thankyou {name.upper()} ,Please Visit Again")
            break
        else:
            print("You have entered a wrong input!!")
    wb.save('account_details_updated.xlsx')
    print("Data updated")




